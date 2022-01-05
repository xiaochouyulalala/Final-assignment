import openpyxl
import os

def make_type1(data,sheet):
    data_list = []
    row1 = data.max_row-1
    for i in range(row1):
        data_list.append((data.cell(i+2,2).value,data.cell(i+2, 4).value))
    data_list=sorted(data_list,key=lambda x:x[1],reverse=1)
    for i in range(20):
        sheet.cell(i+2, 1).value = i+1
        sheet.cell(i+2, 2).value = data_list[i][0]
        sheet.cell(i+2, 3).value = data_list[i][1]
    return sheet

def make_type2(data,cate,sheet):
    data_list = []
    cate_list = {}
    num = {}
    row1 = data.max_row-1
    row2 = cate.max_row-1
    for i in range(row1):
        data_list.append((data.cell(i+2,2).value,data.cell(i+2, 4).value))
    for i in range(row2):
        cate_list[cate.cell(i+2,2).value] = cate.cell(i+2,3).value
        num[cate.cell(i+2,3).value] = 0
    #print(cate_list)
    for each in data_list:
        #print (each)
        if(each[0]):
            each_kind = cate_list[each[0]]
            number = int(num[each_kind]) +int(each[1])
            num[each_kind] = number
    ans=sorted(num.items(),key=lambda x:x[1],reverse=1)
    #print(num)
    for i in range(10):
        sheet.cell(i+2, 1).value = i+1
        sheet.cell(i+2, 2).value = ans[i][0]
        sheet.cell(i+2, 3).value = ans[i][1]
    return sheet

def make_type3(data,sheet):
    data_list = []
    row1 = data.max_row-1
    for i in range(row1):
        data_list.append((data.cell(i+2,2).value,data.cell(i+2, 5).value))
    data_list=sorted(data_list,key=lambda x:x[1],reverse=1)
    for i in range(10):
        sheet.cell(i+2, 1).value = i+1
        sheet.cell(i+2, 2).value = data_list[i][0]
        sheet.cell(i+2, 3).value = data_list[i][1]
    return sheet

def make_type4(data,cate,sheet):
    data_list = []
    cate_list = {}
    num = {}
    row1 = data.max_row-1
    row2 = cate.max_row-1
    for i in range(row1):
        data_list.append((data.cell(i+2,2).value,data.cell(i+2, 5).value))
    for i in range(row2):
        cate_list[cate.cell(i+2,2).value] = cate.cell(i+2,3).value
        num[cate.cell(i+2,3).value] = 0
    #print(cate_list)
    for each in data_list:
        #print (each)
        if(each[0]):
            each_kind = cate_list[each[0]]
            number = int(num[each_kind]) +int(each[1])
            num[each_kind] = number
    ans=sorted(num.items(),key=lambda x:x[1],reverse=1)
    #print(num)
    for i in range(10):
        sheet.cell(i+2, 1).value = i+1
        sheet.cell(i+2, 2).value = ans[i][0]
        sheet.cell(i+2, 3).value = ans[i][1]
    return sheet

def make_type5(data1,data2,data3,sheet):
    data_list = []
    list1 = []
    list2 = []
    list3 = []
    ans = []
    num1 = {}
    num2 = {}
    num3 = {}
    row1 = data1.max_row-1
    for i in range(row1):
        list1.append(data1.cell(i+2,2).value)
        num1[data1.cell(i+2,2).value] = (data1.cell(i+2,4).value,1)
        data_list.append(data1.cell(i+2,2).value)
    row1 = data2.max_row-1
    for i in range(row1):
        list2.append(data2.cell(i+2,2).value)
        num2[data2.cell(i+2,2).value] = (data2.cell(i+2,4).value,1)
        data_list.append(data2.cell(i+2,2).value)
    row1 = data3.max_row-1
    for i in range(row1):
        list3.append(data3.cell(i+2,2).value)
        num3[data3.cell(i+2,2).value] = (data3.cell(i+2,4).value,data3.cell(i+2,5).value)
        data_list.append(data3.cell(i+2,2).value)
    data_list = list(set(data_list))
    list1 = list(set(list1))
    list2 = list(set(list2))
    list3 = list(set(list3))
    for each in data_list:
        number = list1.count(each)+list2.count(each)+list3.count(each)
        #print(number)
        if(number==3):
            #print("YES")
            ans.append((each,num1[each][0]+num2[each][0]+num3[each][0],num1[each][1]+num2[each][1]+num3[each][1]))
    ans=sorted(ans,key=lambda x:(x[1],x[2]),reverse=1)
    #print(ans)
    l = min(10,len(ans))
    for i in range(l):
        sheet.cell(i+2, 1).value = i+1
        sheet.cell(i+2, 2).value = ans[i][0]
        sheet.cell(i+2, 3).value = ans[i][1]
        sheet.cell(i+2, 4).value = ans[i][2]
    return sheet









path = os.path.abspath("datawork.py")
path = os.path.abspath(os.path.dirname(path)+os.path.sep+'.')
fs = os.listdir(path)
xls = openpyxl.Workbook()
xls.create_sheet("北邮TOP20")
xls.create_sheet("西电TOP20")
xls.create_sheet("成电TOP20")
xls.create_sheet("北邮雇主TOP10")
xls.create_sheet("西电雇主TOP10")
xls.create_sheet("成电雇主TOP10")
xls.create_sheet("北邮职位数量TOP10")
xls.create_sheet("北邮职位数量类型TOP10")
xls.create_sheet("最关注ICT行业TOP10")
xls.remove(xls['Sheet'])
sheet_21 = xls["北邮TOP20"]
sheet_22 = xls["西电TOP20"]
sheet_23 = xls["成电TOP20"]
sheet_24 = xls["北邮雇主TOP10"]
sheet_25 = xls["西电雇主TOP10"]
sheet_26 = xls["成电雇主TOP10"]
sheet_27 = xls["北邮职位数量TOP10"]
sheet_28 = xls["北邮职位数量类型TOP10"]
sheet_29 = xls["最关注ICT行业TOP10"]

title1 = ["序号",	"招聘主题","关注度"]
title2 = ["序号",   "雇主类型","关注度"]
title3 = ["序号",	"招聘主题","职位个数"]
title4 = ["序号",   "雇主类型","职位个数"]
title5 = ["序号",   "招聘主题","浏览次数","职位个数"]

for i in range(3):
    sheet_21.cell(1, i+1).value = title1[i]
    sheet_22.cell(1, i+1).value = title1[i]
    sheet_23.cell(1, i+1).value = title1[i]
for i in range(3):
    sheet_24.cell(1, i+1).value = title2[i]
    sheet_25.cell(1, i+1).value = title2[i]
    sheet_26.cell(1, i+1).value = title2[i]
for i in range(3):
    sheet_27.cell(1, i+1).value = title3[i]
for i in range(3):
    sheet_28.cell(1, i+1).value = title4[i]
for i in range(4):
    sheet_29.cell(1, i+1).value = title5[i]

for f in fs:
    if f == 'data.xlsx':
        xl = openpyxl.load_workbook('./'+f)
        break
sheet1 = xl["西电"]
sheet2 = xl["成电"]
sheet3 = xl["北邮"]
sheet4 = xl["分类"]
sheet_21 = make_type1(sheet1,sheet_21)
sheet_22 = make_type1(sheet2,sheet_22)
sheet_23 = make_type1(sheet3,sheet_23)
sheet_24 = make_type2(sheet1,sheet4,sheet_24)
sheet_25 = make_type2(sheet2,sheet4,sheet_25)
sheet_26 = make_type2(sheet3,sheet4,sheet_26)
sheet_27 = make_type3(sheet3,sheet_27)
sheet_28 = make_type4(sheet3,sheet4,sheet_28  )
sheet_29 = make_type5(sheet1,sheet2,sheet3,sheet_29)

xls.save('就业关注分析.xlsx')
xls.close()